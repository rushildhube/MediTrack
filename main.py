import tkinter as tk
from tkinter import filedialog
import firebase_admin
from firebase_admin import credentials, firestore
from kivy.app import App
from kivy.lang import Builder
from openpyxl import Workbook, load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from kivy.uix.textinput import TextInput
from kivy.properties import StringProperty
from kivy.uix.screenmanager import ScreenManager, Screen
import pandas as pd
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from firebase_admin import storage
import smtplib
from datetime import datetime, timedelta


import os
import hashlib
from datetime import datetime
from kivy.core.window import Window


Window.size = (420, 640)

cred = credentials.Certificate("serviceAccountKey.json")
firebase_admin.initialize_app(cred, {'storageBucket': "healthcaresoftwareforpatient.appspot.com"})
db = firestore.client()
# Builder.load_file('main.kv')
currentuser = ''
docs = db.collection('UserData').get()


def add_patient_record(patient_id, record_data):
    try:
        doc_ref = db.collection('PatientsDetails').document(patient_id)
        doc_ref.set(record_data)
        print("Medical record added successfully")
    except Exception as e:
        print(f"Error adding medical record: {e}")


def add_user(user_id, data):
    try:
        doc_ref = db.collection('UserData').document(user_id)
        doc_ref.set(data)
        print("user created successfully")
    except Exception as e:
        print(f"Error adding medical record: {e}")


def hash_password(password):
    """Hash a password for storing."""
    return hashlib.sha256(password.encode()).hexdigest()


def upload_document(patient_id, file_path):
    try:
        # Check if the patient record exists in Firestore
        doc_ref = db.collection('PatientsDetails').document(patient_id)
        doc = doc_ref.get()

        if doc.exists:
            # Generate a reference to the document in Firebase Storage
            bucket = storage.bucket()
            file_name = file_path.split('/')[-1]
            blob = bucket.blob(f"documents/{patient_id}/{file_name}")

            # Upload the document content from the selected file
            blob.upload_from_filename(file_path)

            # Generate a signed URL for temporary access
            expiration_time = timedelta(days=1)  # Use timedelta correctly
            signed_url = blob.generate_signed_url(expiration=expiration_time)

            # Update the patient's record data in Firestore with the signed URL
            doc_ref.update({'document_link': signed_url})

            print(f"Document {file_name} uploaded successfully for patient {patient_id}")
            print(f"Signed URL for the document: {signed_url}")
        else:
            print(f"Patient with ID {patient_id} does not exist.")
    except Exception as e:
        print(f"Error uploading document: {e}")


def share_single_patient_record(patient_id, doctor_email):
    try:
        doc_ref = db.collection('PatientsDetails').document(patient_id)
        patient_data = doc_ref.get().to_dict()

        if patient_data:
            doc_ref.update({'shared_with': firestore.ArrayUnion([doctor_email])})

            # Send email with structured content
            sender_email = 'healthcaredatabasesystem@gmail.com'
            subject = 'Patient Medical Record Shared'
            send_email(sender_email, doctor_email, subject, patient_id, patient_data)

            print(f"Medical record of patient {patient_id} shared with {doctor_email}. Email sent successfully.")
        else:
            print(f"Patient record with ID {patient_id} not found")
    except Exception as e:
        print(f"Error sharing medical record: {e}")


def send_email(sender_email, receiver_email, subject, patient_id, patient_data):
    try:
        smtp_server = 'smtp.gmail.com'
        port = 587
        username = 'healthcaredatabasesystem@gmail.com'
        password = 'monccsrwuupjzsdm'

        body = f"Dear Doctor,\n\nPatient with ID {patient_id} has shared their medical record with you.\n\nMedical Record Details:\n"
        for key, value in patient_data.items():
            body += f"- {key}: {value}\n"

        body += "\nPlease review the details and provide necessary assistance.\n\nBest regards,\nHealthcare System"

        message = MIMEMultipart()
        message['From'] = sender_email
        message['To'] = receiver_email
        message['Subject'] = subject
        message.attach(MIMEText(body, 'plain'))

        server = smtplib.SMTP(smtp_server, port)
        server.starttls()
        server.login(username, password)
        server.send_message(message)
        server.quit()

        print(f"Email sent successfully to {receiver_email}")
    except Exception as e:
        print(f"Error sending email: {e}")


class RestrictedTextInput(TextInput):
    input_restriction = StringProperty('none')  # Set a default value

    def __init__(self, **kwargs):
        super(RestrictedTextInput, self).__init__(**kwargs)
        self.input_restriction = kwargs.get('input_restriction', 'none')

        if self.input_restriction != 'none':
            self.input_filter = self.validate_input


    def validate_input(self, text, from_undo=False):
        if self.input_restriction == 'numeric':
            if text.isdigit() or text == '':
                return text
            else:
                return ''
        elif self.input_restriction == 'alpha':
            if text.isalpha() or text == '':
                return text
            else:
                return ''
        else:
            return text


class HomePage(Screen):
    pass


class LoginScreen(Screen):
    def login(self):
        username = self.ids.login_username.text
        global currentuser
        currentuser = username
        print(currentuser)
        password = self.ids.login_password.text
        hashed_password = hash_password(password)
        doc = db.collection('UserData').document(username)
        doc = doc.get()
        if doc.exists:
            doc = doc.to_dict()
            if doc['password'] == hashed_password:
                currentuser = username
                self.manager.current = 'userdashboard'  # Redirect to UserDashboard
                self.ids.login_label.text = ''
            else:
                self.ids.login_label.text = 'Wrong password.'
        else:
            self.ids.login_label.text = 'Wrong username'


class SignUpScreen(Screen):
    def signup(self):
        username = self.ids.signup_username.text
        password = self.ids.signup_password.text

        if not password:
            self.ids.signup_label.text = 'Password cannot be empty.'
            return

        hashed_password = hash_password(password)

        for f in docs:
            x = f.to_dict()
            if (x['username'] == username):
                print(x['username'])
                self.ids.signup_label.text = 'Username already exists.'
                return

        self.manager.current = 'profile'
        self.ids.signup_label.text = 'Account created, please update latest profile.'
        global currentuser
        currentuser = username
        record_data = {
            'created_at': datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            'username': username,
            'password':hashed_password,
            'name': None,
            'age': None,
            'gender': None,
            'medical_history': None,
            'disease': None,
            'medication': None
        }
        data = {'username':username,
                'password':hashed_password}
        add_user(username,data)
        add_patient_record(username, record_data)


class ProfileScreen(Screen):
    def save_profile(self):
        name = self.ids.profile_name.text
        diagnosis = self.ids.profile_diagnosis.text
        age = self.ids.profile_age.text
        medical_history = self.ids.profile_medical_history.text
        gender = self.ids.profile_gender.text
        medication = self.ids.profile_medication.text
        record_data = {
            'created_at': datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            'name': name,
            'age': age,
            'gender': gender,
            'medical_history': medical_history,
            'disease': diagnosis,
            'medication': medication
        }
        global currentuser

        add_patient_record(currentuser, record_data)
        self.manager.current = 'userdashboard'


class UserDashboard(Screen):
    def select_and_upload_file(self):
        root = tk.Tk()
        root.withdraw()  # Hide the root window

        file_path = filedialog.askopenfilename(initialdir='C:\\', title='Select a file to upload')

        if file_path:
            patient_id = currentuser
            upload_document(patient_id, file_path)
        else:
            print("No file selected.")


class BloodTestScreen(Screen):
    def call(self):
        global currentuser
        filepath = 'records.xlsx'
        if filepath:
            upload_document(currentuser,filepath)


class NewRecordScreen(Screen):
    def save_new_record(self, test_date, blood_group, wbc_count, platelet_count, hemoglobin):
        # Define the filename
        filename = "records.xlsx"

        # Check if the file exists and load it, otherwise create a new workbook
        if os.path.exists(filename):
            try:
                wb = load_workbook(filename)
                ws = wb.active
            except InvalidFileException:
                print("Error: Failed to load existing workbook. Creating a new one.")
                wb = Workbook()
                ws = wb.active
                ws.append(["Test Date", "Blood Group", "WBC Count", "Platelet Count", "Hemoglobin"])
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = "Reports"
            ws.append(["Test Date", "Blood Group", "WBC Count", "Platelet Count", "Hemoglobin"])

        # Check if all fields are filled
        if not all([test_date, blood_group, wbc_count, platelet_count, hemoglobin]):
            print("Error: All fields must be filled")
            return

        # Append the new record
        ws.append([test_date, blood_group, wbc_count, platelet_count, hemoglobin])

        # Save the workbook
        wb.save(filename)
        self.manager.current = "bloodtest"
        print("New record saved successfully.")


class ShowRecordScreen(Screen):
    def open_excel(self, file_path):
        try:

            df = pd.read_excel(file_path)
            excel_data = df.to_string(index=False)  # Convert DataFrame to a string
            self.ids.excel_data.text = excel_data
        except FileNotFoundError:
            self.ids.excel_data.text = "Excel file not found. Please check the file path."
        except Exception as e:
            self.ids.excel_data.text = f"An error occurred: {e}"


class Share(Screen):
    def call(self,email):
        global currentuser
        share_single_patient_record(currentuser,email)


class ScreenManagement(ScreenManager):
    pass


class MobileApp(App):
    def build(self):
        self.title = 'MediTrack'
        return Builder.load_file('main.kv')




if __name__ == '__main__':
    MobileApp().run()
